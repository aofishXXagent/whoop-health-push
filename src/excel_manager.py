"""Excel 管理：从 SQLite 全量重建 4 sheets + 飞书 Drive 上传。"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src import config
from src.database import get_all_daily, get_all_workouts


# ── 样式 ─────────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")

SHEET_DEFS = {
    "Recovery": {
        "columns": [
            ("日期", "date"),
            ("恢复分 (%)", "recovery_score"),
            ("HRV (ms)", "hrv"),
            ("静息心率 (bpm)", "resting_hr"),
            ("血氧 (%)", "spo2"),
            ("皮肤温度偏移 (°C)", "skin_temp"),
        ],
        "source": "daily",
    },
    "Sleep": {
        "columns": [
            ("日期", "date"),
            ("睡眠总时长 (min)", "sleep_total_min"),
            ("深睡 (min)", "sleep_deep_min"),
            ("REM (min)", "sleep_rem_min"),
            ("浅睡 (min)", "sleep_light_min"),
            ("清醒 (min)", "sleep_awake_min"),
            ("周期数", "sleep_cycle_count"),
            ("干扰次数", "disturbance_count"),
            ("睡眠表现 (%)", "sleep_performance"),
            ("睡眠一致性 (%)", "sleep_consistency"),
            ("睡眠效率 (%)", "sleep_efficiency"),
            ("呼吸频率 (/min)", "respiratory_rate"),
            ("基线需求 (min)", "sleep_need_baseline_min"),
            ("睡眠债 (min)", "sleep_need_debt_min"),
            ("压力补偿 (min)", "sleep_need_strain_min"),
        ],
        "source": "daily",
    },
    "Strain": {
        "columns": [
            ("日期", "date"),
            ("压力分", "strain"),
            ("平均心率 (bpm)", "avg_hr"),
            ("最大心率 (bpm)", "max_hr"),
            ("能量消耗 (kJ)", "kilojoules"),
        ],
        "source": "daily",
    },
    "Workouts": {
        "columns": [
            ("日期", "date"),
            ("运动类型", "sport_name"),
            ("压力分", "strain"),
            ("平均心率 (bpm)", "avg_hr"),
            ("最大心率 (bpm)", "max_hr"),
            ("距离 (m)", "distance_m"),
            ("海拔增益 (m)", "altitude_gain_m"),
            ("时长 (min)", "duration_min"),
            ("能量 (kJ)", "kilojoules"),
            ("Z0 休息 (min)", "zone_0_min"),
            ("Z1 热身 (min)", "zone_1_min"),
            ("Z2 有氧 (min)", "zone_2_min"),
            ("Z3 阈值 (min)", "zone_3_min"),
            ("Z4 无氧 (min)", "zone_4_min"),
            ("Z5 极限 (min)", "zone_5_min"),
        ],
        "source": "workouts",
    },
}


def _apply_header_style(ws, num_cols: int):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def _fmt(v):
    if v is None:
        return "暂无"
    if isinstance(v, float):
        return round(v, 1)
    return v


def rebuild_excel(excel_path=None):
    """从 SQLite 全量重建 Excel 文件（4 sheets）。"""
    excel_path = excel_path or config.EXCEL_PATH

    daily_rows = get_all_daily()
    workout_rows = get_all_workouts()
    source_map = {"daily": daily_rows, "workouts": workout_rows}

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for sheet_name, sheet_def in SHEET_DEFS.items():
        ws = wb.create_sheet(title=sheet_name)
        columns = sheet_def["columns"]
        rows = source_map[sheet_def["source"]]

        # 写表头
        for col_idx, (header, _) in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写数据
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, (_, key) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=_fmt(row_data.get(key)))

        _apply_header_style(ws, len(columns))
        _auto_width(ws)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    print(f"[Excel] Saved to {excel_path} ({len(daily_rows)} daily + {len(workout_rows)} workouts)")
